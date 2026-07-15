from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from hypothesis_factory.data_loaders.real_case_loader import TailingsObservation


SIZE_CLASSES = ["+125", "+71", "-125+71", "-71+45", "-45+20", "-20+10", "-10"]


@dataclass
class TailingsParseResult:
    source_file: str
    factory: str
    observations: list[TailingsObservation] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _safe_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and value.startswith("#"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _row_values(sheet, row: int) -> list[Any]:
    return [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]


def parse_tailings_xlsx(path: Path, factory: str) -> TailingsParseResult:
    """Best-effort openpyxl parser for the case spreadsheets.

    The real workbooks are semi-structured, so the parser anchors on text labels
    in column B and collects nearby size-class rows instead of relying on fixed
    coordinates.
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return TailingsParseResult(str(path), factory, warnings=[f"openpyxl unavailable: {exc}"])

    result = TailingsParseResult(source_file=str(path), factory=factory)
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        result.warnings.append(f"cannot read workbook: {exc}")
        return result

    sheet = workbook[workbook.sheetnames[0]]
    current_tailings_type = "Хвосты отвальные"
    for row in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row=row, column=2).value or "").strip()
        label_lower = label.lower()
        if "хвост" in label_lower:
            current_tailings_type = label
        if label not in SIZE_CLASSES:
            continue
        values = _row_values(sheet, row)
        numeric = [_safe_number(value) for value in values]
        if any(isinstance(value, str) and value.startswith("#") for value in values):
            result.warnings.append(f"{path.name}: ignored formula error at row {row}")
        numbers = [value for value in numeric if value is not None]
        if not numbers:
            continue
        class_share = numbers[0] if len(numbers) >= 1 else None
        element_28_share = numbers[1] if len(numbers) >= 2 else None
        element_28_t = numbers[2] if len(numbers) >= 3 else None
        element_29_share = numbers[3] if len(numbers) >= 4 else None
        element_29_t = numbers[4] if len(numbers) >= 5 else None
        for element, share, mass in [
            ("element_28", element_28_share, element_28_t),
            ("element_29", element_29_share, element_29_t),
        ]:
            result.observations.append(
                TailingsObservation(
                    source_file=str(path),
                    factory=factory,
                    tailings_type=current_tailings_type,
                    element=element,
                    particle_size_class=label,
                    loss_share_pct=share,
                    loss_mass_t=mass,
                    class_share_pct=class_share,
                    row_ref=f"row {row}",
                )
            )
    return result
